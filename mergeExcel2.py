import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side

def merge_excel_files(excel_files):
    combined_dataframes = []
    for file in excel_files:
        data = pd.read_excel(file, sheet_name=None)
        for sheet_name, df in data.items():
            combined_dataframes.append(df)
    combined_df = pd.concat(combined_dataframes, ignore_index=True)
    return combined_df

def save_excel_with_separator(df, filename):
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    df.to_excel(writer, index=False, header=True, startrow=0, startcol=0, sheet_name='Sheet1')
    writer.save()

    # Adding a bold line between sheets
    wb = Workbook()
    ws = wb.active

    # Accessing the workbook and worksheet
    wb = openpyxl.load_workbook(filename)
    worksheet = wb.active

    # Adding a bold line between sheets
    border = Border(bottom=Side(style='medium'))
    for row_num in range(len(df) + 1, len(df) + 3):
        for col_num in range(len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
    
    wb.save(filename)

def main():
    st.title("Excel File Merger")

    st.write("Upload Excel Files")
    uploaded_files = st.file_uploader("Choose Excel files", accept_multiple_files=True, type="xlsx")

    if uploaded_files:
        st.write("Files uploaded successfully!")
        st.write("Merging files...")

        combined_df = merge_excel_files(uploaded_files)

        st.write("Merged Data:")
        st.write(combined_df)

        st.write("Save Merged Excel File")
        file_name = st.text_input("Enter the filename for the merged Excel file (without extension):")
        if file_name:
            save_excel_with_separator(combined_df, f"{file_name}.xlsx")
            st.success(f"File '{file_name}.xlsx' saved successfully!")

